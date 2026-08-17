# -*- coding: utf-8 -*-
"""
Sistema de extracción de comprobantes de pago con IA — versión 1
=================================================================
Usa Google Gemini Flash (API gratuita, sin tarjeta de crédito).
Flujo: lee imágenes de "entrada", envía cada una a Gemini, extrae los datos
en JSON, valida, escribe en salida/comprobantes.xlsx y mueve la imagen.

Uso normal: doble clic en PROCESAR.bat
"""

import io
import json
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Verificación de dependencias
# ---------------------------------------------------------------------------
try:
    from google import genai
    from google.genai import types
except ImportError:
    print("ERROR: Faltan las dependencias del sistema.")
    print("Solucion: haz doble clic en INSTALAR.bat y espera a que termine.")
    input("\nPresiona ENTER para cerrar...")
    sys.exit(1)

try:
    from PIL import Image
except ImportError:
    print("ERROR: Faltan las dependencias del sistema.")
    print("Solucion: haz doble clic en INSTALAR.bat y espera a que termine.")
    input("\nPresiona ENTER para cerrar...")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Faltan las dependencias del sistema (openpyxl).")
    print("Solucion: abre una terminal y ejecuta: pip install openpyxl")
    input("\nPresiona ENTER para cerrar...")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuración general
# ---------------------------------------------------------------------------
BASE = Path(__file__).resolve().parent
CARPETA_ENTRADA = BASE / "entrada"
CARPETA_PROCESADOS = BASE / "procesados"
CARPETA_REVISION = BASE / "revision_manual"
CARPETA_SALIDA = BASE / "salida"
ARCHIVO_XLSX = CARPETA_SALIDA / "comprobantes.xlsx"
ARCHIVO_CONFIG = BASE / "config.txt"

# Modelo de Gemini Flash (gratuito, 1500 solicitudes/día, 15/minuto)
MODELO = "gemini-3.6-flash"

LADO_MAXIMO = 1568  # px en el lado más largo
EXTENSIONES = {".jpg", ".jpeg", ".png", ".webp"}

# Pausa entre solicitudes para respetar el límite de 15 por minuto
PAUSA_SEGUNDOS = 4.5

COLUMNAS = [
    "fecha_procesado",
    "archivo_origen",
    "banco_app",
    "numero_comprobante",
    "numero_cuenta",
    "nombre_cliente",
    "valor_pago",
    "fecha_pago",
    "estado",
]

PROMPT = """Analiza esta imagen de un comprobante de pago colombiano (Nequi, Bancolombia o Daviplata).

Extrae los datos y responde ÚNICAMENTE con un objeto JSON válido, sin texto adicional y sin marcas de código:

{
  "banco_app": "Nequi" | "Bancolombia" | "Daviplata" | "Otro",
  "numero_comprobante": "solo dígitos, o null",
  "numero_cuenta": "dígitos; si está parcialmente oculto transcribe lo visible; o null",
  "nombre_cliente": "nombre tal como aparece, o null",
  "valor_pago": entero en pesos, o null,
  "fecha_pago": "AAAA-MM-DD, o null"
}

Guía por aplicación:
- Nequi: el número de comprobante aparece como "Referencia". La cuenta suele ser un número de celular de 10 dígitos que inicia en 3.
- Bancolombia: el número de comprobante aparece como "Comprobante No." o "Número de aprobación".
- Daviplata: el número de comprobante aparece como "No. de aprobación". La cuenta suele ser un celular de 10 dígitos.

Reglas estrictas:
1. NUNCA inventes ni completes datos. Si un dato no es legible o no aparece en la imagen, usa null.
2. numero_comprobante y numero_cuenta: solo dígitos, sin espacios, puntos ni guiones. Si la cuenta aparece parcialmente oculta (por ejemplo ***1234), transcribe exactamente lo visible incluyendo los asteriscos.
3. nombre_cliente: transcribe el nombre de la persona exactamente como aparece, incluso si está parcialmente oculto (por ejemplo "MARIA C***"). No lo completes ni lo adivines.
4. valor_pago: número entero en pesos colombianos, sin puntos, comas ni símbolo $. Ejemplo: 150000.
5. fecha_pago: la fecha de la transacción en formato AAAA-MM-DD. Si el año no es visible, usa null."""


# ---------------------------------------------------------------------------
# Funciones auxiliares
# ---------------------------------------------------------------------------
def leer_api_key():
    """Lee la clave de API desde config.txt."""
    if not ARCHIVO_CONFIG.exists():
        print("ERROR: No se encontro el archivo config.txt.")
        print("Debe estar en la misma carpeta que este programa.")
        return None
    for linea in ARCHIVO_CONFIG.read_text(encoding="utf-8-sig").splitlines():
        linea = linea.strip()
        if linea.startswith("#") or "=" not in linea:
            continue
        clave, valor = linea.split("=", 1)
        if clave.strip() == "API_KEY":
            valor = valor.strip()
            if not valor or "PEGA_AQUI" in valor:
                print("ERROR: Falta configurar la clave de API.")
                print("Abre config.txt con el Bloc de notas y pega tu clave")
                print("despues de API_KEY= (la clave empieza con AIza...).")
                print("Se obtiene gratis en https://aistudio.google.com/apikey")
                return None
            return valor
    print("ERROR: config.txt no contiene la linea API_KEY=.")
    return None


def crear_carpetas():
    for carpeta in (CARPETA_ENTRADA, CARPETA_PROCESADOS, CARPETA_REVISION, CARPETA_SALIDA):
        carpeta.mkdir(exist_ok=True)


def listar_imagenes():
    return sorted(
        p for p in CARPETA_ENTRADA.iterdir()
        if p.is_file() and p.suffix.lower() in EXTENSIONES
    )


def cargar_comprobantes_existentes():
    """Lee el Excel existente y devuelve el conjunto de comprobantes ya registrados."""
    existentes = set()
    if ARCHIVO_XLSX.exists():
        libro = load_workbook(ARCHIVO_XLSX, read_only=True)
        hoja = libro.active
        indice = COLUMNAS.index("numero_comprobante")
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if fila and len(fila) > indice and fila[indice] is not None:
                numero = str(fila[indice]).strip()
                if numero:
                    existentes.add(numero)
        libro.close()
    return existentes


# ---------------------------------------------------------------------------
# Escritura del archivo Excel
# ---------------------------------------------------------------------------
RELLENO_ENCABEZADO = PatternFill("solid", fgColor="16213E")
FUENTE_ENCABEZADO = Font(bold=True, color="FFFFFF")
RELLENO_OK = PatternFill("solid", fgColor="C8E6C9")       # verde claro
RELLENO_REVISAR = PatternFill("solid", fgColor="FFF9C4")  # amarillo claro


def _ajustar_anchos(hoja, valores):
    """Ensancha cada columna si el valor nuevo es más largo que el ancho actual."""
    for indice, valor in enumerate(valores, start=1):
        letra = get_column_letter(indice)
        ancho = len(str(valor)) + 3
        actual = hoja.column_dimensions[letra].width or 0
        if ancho > actual:
            hoja.column_dimensions[letra].width = ancho


def abrir_libro():
    """Abre salida/comprobantes.xlsx, o lo crea con los encabezados formateados."""
    if ARCHIVO_XLSX.exists():
        libro = load_workbook(ARCHIVO_XLSX)
        return libro, libro.active
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Comprobantes"
    hoja.append(COLUMNAS)
    for celda in hoja[1]:
        celda.font = FUENTE_ENCABEZADO
        celda.fill = RELLENO_ENCABEZADO
    _ajustar_anchos(hoja, COLUMNAS)
    return libro, hoja


def agregar_fila(hoja, fila):
    """Agrega la fila (dict) al final de la hoja y aplica los formatos."""
    valores = [fila[columna] for columna in COLUMNAS]
    hoja.append(valores)
    numero_fila = hoja.max_row

    celda_valor = hoja.cell(row=numero_fila, column=COLUMNAS.index("valor_pago") + 1)
    celda_valor.number_format = "#,##0"

    celda_estado = hoja.cell(row=numero_fila, column=COLUMNAS.index("estado") + 1)
    if fila["estado"] == "OK":
        celda_estado.fill = RELLENO_OK
    else:
        celda_estado.fill = RELLENO_REVISAR

    _ajustar_anchos(hoja, valores)


def preparar_imagen(ruta):
    """Abre la imagen, la convierte a RGB, la reduce a 1568px máximo.
    Devuelve los bytes JPEG y el mime type."""
    img = Image.open(ruta)
    if img.mode != "RGB":
        img = img.convert("RGB")
    if max(img.size) > LADO_MAXIMO:
        img.thumbnail((LADO_MAXIMO, LADO_MAXIMO), Image.LANCZOS)
    buffer = io.BytesIO()
    img.save(buffer, format="JPEG", quality=85)
    return buffer.getvalue()


def extraer_datos(cliente, imagen_bytes):
    """Envía la imagen a Gemini y devuelve el diccionario con los datos."""
    respuesta = cliente.models.generate_content(
        model=MODELO,
        contents=[
            types.Content(
                role="user",
                parts=[
                    types.Part.from_bytes(
                        data=imagen_bytes,
                        mime_type="image/jpeg",
                    ),
                    types.Part.from_text(text=PROMPT),
                ],
            )
        ],
        config=types.GenerateContentConfig(
            temperature=0,
            # El modelo usa tokens de razonamiento que cuentan contra este
            # límite; 600 no alcanza y la respuesta sale cortada (MAX_TOKENS).
            max_output_tokens=2000,
            response_mime_type="application/json",
        ),
    )
    texto = respuesta.text or ""
    if not texto.strip():
        raise ValueError("El modelo devolvio una respuesta vacia.")
    return parsear_json(texto)


def parsear_json(texto):
    """Extrae el objeto JSON de la respuesta, tolerando marcas de código."""
    texto = texto.strip()
    # Quitar marcas ```json ... ```
    texto = re.sub(r"^```(?:json)?\s*", "", texto)
    texto = re.sub(r"\s*```$", "", texto)
    inicio, fin = texto.find("{"), texto.rfind("}")
    if inicio == -1 or fin == -1:
        raise ValueError("La respuesta del modelo no contiene JSON.")
    return json.loads(texto[inicio : fin + 1])


def limpiar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def normalizar_valor(valor):
    """Convierte el valor del pago a entero, tolerando formatos como '150.000'."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    digitos = re.sub(r"[^\d]", "", str(valor))
    return int(digitos) if digitos else None


def fecha_valida(texto):
    try:
        datetime.strptime(texto, "%Y-%m-%d")
        return True
    except (ValueError, TypeError):
        return False


def validar(datos, comprobantes_existentes):
    """Aplica validaciones. Devuelve (estado, lista de motivos)."""
    motivos = []
    comprobante = limpiar_texto(datos.get("numero_comprobante"))
    cuenta = limpiar_texto(datos.get("numero_cuenta"))
    nombre = limpiar_texto(datos.get("nombre_cliente"))
    valor = normalizar_valor(datos.get("valor_pago"))
    fecha = limpiar_texto(datos.get("fecha_pago"))

    if not comprobante:
        motivos.append("falta el numero de comprobante")
    elif comprobante in comprobantes_existentes:
        motivos.append("comprobante ya registrado antes (posible pago duplicado)")

    if not cuenta:
        motivos.append("falta el numero de cuenta")
    elif not cuenta.isdigit():
        motivos.append("cuenta con caracteres no numericos (posible dato oculto)")

    if not nombre:
        motivos.append("falta el nombre del cliente")
    elif "*" in nombre:
        motivos.append("nombre parcialmente oculto en el comprobante")

    if valor is None or valor <= 0:
        motivos.append("falta el valor del pago o no es valido")

    if not fecha:
        motivos.append("falta la fecha del pago")
    elif not fecha_valida(fecha):
        motivos.append("la fecha no tiene el formato esperado")

    estado = "OK" if not motivos else "REVISAR"
    return estado, motivos


def construir_fila(archivo, datos, estado):
    valor = normalizar_valor(datos.get("valor_pago"))
    return {
        "fecha_procesado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "archivo_origen": archivo.name,
        "banco_app": limpiar_texto(datos.get("banco_app")),
        "numero_comprobante": limpiar_texto(datos.get("numero_comprobante")),
        "numero_cuenta": limpiar_texto(datos.get("numero_cuenta")),
        "nombre_cliente": limpiar_texto(datos.get("nombre_cliente")),
        "valor_pago": valor if valor is not None else "",
        "fecha_pago": limpiar_texto(datos.get("fecha_pago")),
        "estado": estado,
    }


def mover_imagen(archivo, destino_dir):
    destino = destino_dir / archivo.name
    if destino.exists():
        marca = datetime.now().strftime("%Y%m%d%H%M%S")
        destino = destino_dir / f"{archivo.stem}_{marca}{archivo.suffix}"
    shutil.move(str(archivo), str(destino))


# ---------------------------------------------------------------------------
# Programa principal
# ---------------------------------------------------------------------------
def main():
    print("=" * 58)
    print("  SISTEMA DE EXTRACCION DE COMPROBANTES — version 1")
    print("  Motor: Google Gemini Flash (gratuito)")
    print("=" * 58)
    print()

    crear_carpetas()

    clave = leer_api_key()
    if not clave:
        input("\nPresiona ENTER para cerrar...")
        return

    imagenes = listar_imagenes()
    if not imagenes:
        print("No hay imagenes para procesar.")
        print('Guarda los comprobantes (jpg, png o webp) en la carpeta "entrada"')
        print("y vuelve a hacer doble clic en PROCESAR.bat.")
        input("\nPresiona ENTER para cerrar...")
        return

    total = len(imagenes)
    print(f"Se encontraron {total} imagenes para procesar.\n")

    if total > 1500:
        print("ADVERTENCIA: Gemini gratis permite hasta 1,500 solicitudes por dia.")
        print(f"Tienes {total} imagenes. Se procesaran las primeras 1,500 hoy.\n")

    comprobantes_existentes = cargar_comprobantes_existentes()
    cliente = genai.Client(api_key=clave)

    total_ok = total_revisar = total_error = 0
    libro, hoja = abrir_libro()

    for indice, archivo in enumerate(imagenes, start=1):
        if indice > 1500:
            print(f"\nSe alcanzo el limite diario de 1,500. Continua manana.")
            break

        print(f"[{indice}/{total}] {archivo.name} ... ", end="", flush=True)
        try:
            imagen_bytes = preparar_imagen(archivo)
            datos = extraer_datos(cliente, imagen_bytes)
            estado, motivos = validar(datos, comprobantes_existentes)
            fila = construir_fila(archivo, datos, estado)
            agregar_fila(hoja, fila)
            libro.save(ARCHIVO_XLSX)

            if fila["numero_comprobante"]:
                comprobantes_existentes.add(fila["numero_comprobante"])

            if estado == "OK":
                total_ok += 1
                mover_imagen(archivo, CARPETA_PROCESADOS)
                print("[OK]")
            else:
                total_revisar += 1
                mover_imagen(archivo, CARPETA_REVISION)
                print("[REVISAR] -> " + "; ".join(motivos))

            # Pausa para respetar límite de 15 solicitudes/minuto
            if indice < total:
                time.sleep(PAUSA_SEGUNDOS)

        except Exception as error:
            error_str = str(error).lower()
            if "api key" in error_str or "authenticate" in error_str or "permission" in error_str:
                print("[ERROR]")
                print("\nLa clave de API no es valida. Revisa config.txt.")
                print("La clave se obtiene gratis en https://aistudio.google.com/apikey")
                break
            elif "429" in error_str or "resource" in error_str or "quota" in error_str:
                print("[ESPERA]")
                print("   Se alcanzo el limite por minuto. Esperando 60 segundos...")
                time.sleep(60)
                # Reintentar esta imagen
                try:
                    imagen_bytes = preparar_imagen(archivo)
                    datos = extraer_datos(cliente, imagen_bytes)
                    estado, motivos = validar(datos, comprobantes_existentes)
                    fila = construir_fila(archivo, datos, estado)
                    agregar_fila(hoja, fila)
                    libro.save(ARCHIVO_XLSX)
                    if fila["numero_comprobante"]:
                        comprobantes_existentes.add(fila["numero_comprobante"])
                    if estado == "OK":
                        total_ok += 1
                        mover_imagen(archivo, CARPETA_PROCESADOS)
                        print(f"   Reintento {archivo.name} ... [OK]")
                    else:
                        total_revisar += 1
                        mover_imagen(archivo, CARPETA_REVISION)
                        print(f"   Reintento {archivo.name} ... [REVISAR]")
                except Exception:
                    total_error += 1
                    print(f"   Reintento fallido. La imagen queda en 'entrada'.")
            else:
                total_error += 1
                print(f"[ERROR] -> {error}")
                print("   La imagen se queda en 'entrada' para reintentar luego.")

    print()
    print("-" * 58)
    print("RESUMEN")
    print(f"  Procesados correctamente (OK): {total_ok}")
    print(f"  Para revision manual:          {total_revisar}")
    print(f"  Con error (quedan en entrada): {total_error}")
    print()
    print(f"Resultados guardados en: salida\\comprobantes.xlsx (abrelo con Excel)")
    if total_revisar:
        print('Las imagenes dudosas estan en la carpeta "revision_manual".')
    input("\nPresiona ENTER para cerrar...")


if __name__ == "__main__":
    main()
